"""把聚合结果导出到 CSV 与多 sheet 的 xlsx 文件。"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .heroes import HeroIndex
from .stats import FactionStats, HeroStats, MatchRow, PlayerStats


# ---------- 把 dataclass 转成行 ----------


def faction_rows(f: FactionStats) -> list[list[Any]]:
    return [
        ["阵营", "胜场", "胜率"],
        ["天辉 Radiant", f.radiant_wins, f"{f.radiant_winrate * 100:.1f}%"],
        ["夜魇 Dire", f.dire_wins, f"{(1 - f.radiant_winrate) * 100:.1f}%" if f.total else "0.0%"],
        ["合计", f.total, ""],
    ]


def player_rows(players: dict[int, PlayerStats], hero_index: HeroIndex) -> list[list[Any]]:
    header = ["account_id", "玩家", "场数", "胜场", "胜率", "K", "D", "A", "KDA", "常用英雄 Top3"]
    rows: list[list[Any]] = [header]
    sorted_players = sorted(players.values(), key=lambda p: (-p.matches, -p.winrate))
    for p in sorted_players:
        top = p.top_heroes(hero_index, 3)
        top_str = "; ".join(f"{name}({cnt}场/{w}胜)" for name, cnt, w in top)
        rows.append(
            [
                p.account_id,
                p.name,
                p.matches,
                p.wins,
                f"{p.winrate * 100:.1f}%",
                p.kills,
                p.deaths,
                p.assists,
                f"{p.kda:.2f}",
                top_str,
            ]
        )
    return rows


def hero_rows(heroes: dict[int, HeroStats], hero_index: HeroIndex) -> list[list[Any]]:
    header = ["英雄", "出场", "胜场", "胜率", "总K", "总D", "总A", "累计KDA"]
    rows: list[list[Any]] = [header]
    sorted_heroes = sorted(heroes.values(), key=lambda h: (-h.picks, -h.winrate))
    for h in sorted_heroes:
        rows.append(
            [
                hero_index.name(h.hero_id),
                h.picks,
                h.wins,
                f"{h.winrate * 100:.1f}%",
                h.kills,
                h.deaths,
                h.assists,
                f"{h.kda:.2f}",
            ]
        )
    return rows


def match_rows(matches: list[MatchRow]) -> list[list[Any]]:
    header = [
        "match_id",
        "开始时间",
        "时长",
        "获胜方",
        "天辉比分",
        "夜魇比分",
        "天辉阵容",
        "夜魇阵容",
    ]
    rows: list[list[Any]] = [header]
    # 按开始时间倒序
    for m in sorted(matches, key=lambda x: x.start_time, reverse=True):
        mm, ss = divmod(m.duration_sec, 60)
        rows.append(
            [
                m.match_id,
                m.start_time.strftime("%Y-%m-%d %H:%M"),
                f"{mm:02d}:{ss:02d}",
                "天辉" if m.radiant_win else "夜魇",
                m.radiant_score,
                m.dire_score,
                m.radiant_players,
                m.dire_players,
            ]
        )
    return rows


# ---------- 写文件 ----------


def write_csv(rows: list[list[Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    # utf-8-sig 让 Excel 直接打开 CSV 时中文不乱码
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def write_xlsx(sheets: dict[str, list[list[Any]]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet 名上限 31 字符
        for row in rows:
            ws.append(row)
        # 表头样式
        if rows:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            # 列宽自适应（粗略：取每列最长字符数 + 2）
            for col_idx in range(1, len(rows[0]) + 1):
                max_len = max(
                    (len(str(r[col_idx - 1])) for r in rows if col_idx - 1 < len(r)),
                    default=10,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
            ws.freeze_panes = "A2"

    wb.save(path)


def export_all(result: dict[str, Any], hero_index: HeroIndex, output_dir: Path, league_id: int) -> Path:
    """把四个维度都写出去：一份 xlsx + 各自的 csv。返回 xlsx 路径。"""
    sheets = {
        "阵营胜率": faction_rows(result["faction"]),
        "玩家战绩": player_rows(result["players"], hero_index),
        "英雄统计": hero_rows(result["heroes"], hero_index),
        "对局列表": match_rows(result["matches"]),
    }
    xlsx_path = output_dir / f"league_{league_id}_stats.xlsx"
    write_xlsx(sheets, xlsx_path)

    csv_dir = output_dir / f"league_{league_id}_csv"
    name_map = {
        "阵营胜率": "faction.csv",
        "玩家战绩": "players.csv",
        "英雄统计": "heroes.csv",
        "对局列表": "matches.csv",
    }
    for sheet_name, rows in sheets.items():
        write_csv(rows, csv_dir / name_map[sheet_name])

    return xlsx_path
